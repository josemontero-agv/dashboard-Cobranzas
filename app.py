# app.py - Dashboard de Ventas Farmacéuticas

from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, jsonify
from dotenv import load_dotenv
from odoo_manager import OdooManager
import os
import pandas as pd
import json
import io
import calendar
from datetime import datetime, timedelta

load_dotenv()
app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY')

# Configuración para deshabilitar cache de templates
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

# --- Inicialización de Managers ---
data_manager = OdooManager()

# Almacenamiento local para metas (reemplaza Google Sheets)
# En producción, esto debería ser una base de datos
LOCAL_STORAGE = {
    'metas_por_linea': {},
    'metas_vendedores': {},
    'equipos': {}
}

# --- Funciones Auxiliares ---

def get_meses_del_año(año):
    """Genera una lista de meses para un año específico."""
    meses_nombres = [
        'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]
    meses_disponibles = []
    for i in range(1, 13):
        mes_key = f"{año}-{i:02d}"
        mes_nombre = f"{meses_nombres[i-1]} {año}"
        meses_disponibles.append({'key': mes_key, 'nombre': mes_nombre})
    return meses_disponibles

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        if data_manager.authenticate_user(username, password):
            session['username'] = username
            flash('¡Inicio de sesión exitoso!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Usuario o contraseña incorrectos.', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Has cerrado sesión correctamente.', 'info')
    return redirect(url_for('login'))

@app.route('/')
@app.route('/sales', methods=['GET', 'POST'])
def sales():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    try:
        # Obtener opciones de filtro
        filter_options = data_manager.get_filter_options()
        
        # Obtener filtros de la request
        selected_filters = {
            'date_from': request.form.get('date_from') or request.args.get('date_from'),
            'date_to': request.form.get('date_to') or request.args.get('date_to'),
            'linea_id': request.form.get('linea_id') or request.args.get('linea_id'),
            'partner_id': request.form.get('partner_id') or request.args.get('partner_id')
        }
        
        # Convertir strings vacíos a None
        for key, value in selected_filters.items():
            if value == '':
                selected_filters[key] = None
            elif key in ['linea_id', 'partner_id'] and value is not None:
                try:
                    selected_filters[key] = int(value)
                except (ValueError, TypeError):
                    selected_filters[key] = None
        
        # Obtener datos
        sales_data = data_manager.get_sales_lines(
            date_from=selected_filters['date_from'],
            date_to=selected_filters['date_to'],
            partner_id=selected_filters['partner_id'],
            linea_id=selected_filters['linea_id'],
            limit=1000
        )
        
        # Filtrar VENTA INTERNACIONAL (exportaciones)
        sales_data_filtered = []
        for sale in sales_data:
            linea_comercial = sale.get('commercial_line_national_id')
            if linea_comercial and isinstance(linea_comercial, list) and len(linea_comercial) > 1:
                nombre_linea = linea_comercial[1].upper()
                if 'VENTA INTERNACIONAL' in nombre_linea:
                    continue
            
            # También filtrar por canal de ventas
            canal_ventas = sale.get('sales_channel_id')
            if canal_ventas and isinstance(canal_ventas, list) and len(canal_ventas) > 1:
                nombre_canal = canal_ventas[1].upper()
                if 'VENTA INTERNACIONAL' in nombre_canal or 'INTERNACIONAL' in nombre_canal:
                    continue
            
            sales_data_filtered.append(sale)
        
        return render_template('sales.html', 
                             sales_data=sales_data_filtered,
                             filter_options=filter_options,
                             selected_filters=selected_filters,
                             fecha_actual=datetime.now())
    
    except Exception as e:
        flash(f'Error al obtener datos: {str(e)}', 'danger')
        return render_template('sales.html', 
                             sales_data=[],
                             filter_options={'lineas': [], 'clientes': []},
                             selected_filters={},
                             fecha_actual=datetime.now())

@app.route('/dashboard', methods=['GET', 'POST'])
def dashboard():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    try:
        # Obtener año actual y mes seleccionado
        fecha_actual = datetime.now()
        año_actual = fecha_actual.year
        mes_seleccionado = request.args.get('mes', fecha_actual.strftime('%Y-%m'))
        
        # Crear todos los meses del año actual
        meses_disponibles = get_meses_del_año(año_actual)
        
        # Obtener nombre del mes seleccionado
        mes_obj = next((m for m in meses_disponibles if m['key'] == mes_seleccionado), None)
        mes_nombre = mes_obj['nombre'] if mes_obj else "Mes Desconocido"
        
        # Obtener día correcto según el mes seleccionado
        if mes_seleccionado == fecha_actual.strftime('%Y-%m'):
            # Mes actual: usar día actual
            dia_actual = fecha_actual.day
        else:
            # Mes pasado: usar último día del mes para mostrar el total del mes completo
            año_sel, mes_sel = mes_seleccionado.split('-')
            ultimo_dia = calendar.monthrange(int(año_sel), int(mes_sel))[1]
            dia_actual = ultimo_dia
        
        # Obtener metas del mes seleccionado desde la sesión
        metas_historicas = LOCAL_STORAGE.get('metas_por_linea', {})
        metas_del_mes = metas_historicas.get(mes_seleccionado, {}).get('metas', {})
        metas_ipn_del_mes = metas_historicas.get(mes_seleccionado, {}).get('metas_ipn', {})
        
        # Líneas comerciales estáticas
        lineas_comerciales_estaticas = [
            {'nombre': 'PETMEDICA', 'id': 'petmedica'},
            {'nombre': 'AGROVET', 'id': 'agrovet'},
            {'nombre': 'PET NUTRISCIENCE', 'id': 'pet_nutriscience'},
            {'nombre': 'AVIVET', 'id': 'avivet'},
            {'nombre': 'ECOMMERCE', 'id': 'ecommerce'},
            {'nombre': 'OTROS', 'id': 'otros'},
            {'nombre': 'GENVET', 'id': 'genvet'},
            {'nombre': 'LICITACIÓN', 'id': 'licitacion'},
            {'nombre': 'Ninguno', 'id': 'ninguno'},
        ]
        
        # Obtener datos reales de ventas desde Odoo
        try:
            # Calcular fechas para el mes seleccionado
            año_sel, mes_sel = mes_seleccionado.split('-')
            fecha_inicio = f"{año_sel}-{mes_sel}-01"
            
            # Último día del mes
            ultimo_dia = calendar.monthrange(int(año_sel), int(mes_sel))[1]
            fecha_fin = f"{año_sel}-{mes_sel}-{ultimo_dia}"
            
            # Obtener datos de ventas reales desde Odoo
            sales_data = data_manager.get_sales_lines(
                date_from=fecha_inicio,
                date_to=fecha_fin,
                limit=5000
            )
            
            print(f"📊 Obtenidas {len(sales_data)} líneas de ventas para el dashboard")
            
        except Exception as e:
            print(f"⚠️ Error obteniendo datos de Odoo: {e}")
            sales_data = []
        
        # Procesar datos de ventas por línea comercial
        datos_lineas = []
        total_meta = 0
        total_venta = 0
        total_meta_pn = 0
        total_venta_pn = 0
        total_vencimiento = 0
        
        # Mapeo de líneas comerciales de Odoo a IDs locales
        mapeo_lineas = {
            'PETMEDICA': 'petmedica',
            'AGROVET': 'agrovet', 
            'PET NUTRISCIENCE': 'pet_nutriscience',
            'AVIVET': 'avivet',
            'ECOMMERCE': 'ecommerce',
            'OTROS': 'otros',
            'GENVET': 'genvet',
            'LICITACIÓN': 'licitacion'
        }
        
        # Calcular ventas reales por línea comercial
        ventas_por_linea = {}
        ventas_por_ruta = {}
        ventas_ipn_por_linea = {} # Nueva variable para ventas de productos nuevos
        ventas_por_producto = {}
        ciclo_vida_por_producto = {}
        ventas_por_ciclo_vida = {}
        ventas_por_forma = {}
        for sale in sales_data:
            # Excluir VENTA INTERNACIONAL (exportaciones)
            linea_comercial = sale.get('commercial_line_national_id')
            nombre_linea_actual = None
            if linea_comercial and isinstance(linea_comercial, list) and len(linea_comercial) > 1:
                nombre_linea_actual = linea_comercial[1].upper()
                if 'VENTA INTERNACIONAL' in nombre_linea_actual:
                    continue
            
            # También filtrar por canal de ventas
            canal_ventas = sale.get('sales_channel_id')
            if canal_ventas and isinstance(canal_ventas, list) and len(canal_ventas) > 1:
                nombre_canal = canal_ventas[1].upper()
                if 'VENTA INTERNACIONAL' in nombre_canal or 'INTERNACIONAL' in nombre_canal:
                    continue
            
            # Procesar el balance de la venta
            balance_float = float(sale.get('balance', 0))
            if balance_float != 0:
                
                # Sumar a ventas totales por línea
                if nombre_linea_actual:
                    ventas_por_linea[nombre_linea_actual] = ventas_por_linea.get(nombre_linea_actual, 0) + balance_float
                
                # LÓGICA FINAL: Sumar si la RUTA (route_id) coincide con los valores especificados
                ruta = sale.get('route_id')
                # Se cambia la comparación al ID de la ruta (ruta[0]) para evitar problemas con traducciones.
                if isinstance(ruta, list) and len(ruta) > 0 and ruta[0] in [18, 19]:
                    if nombre_linea_actual:
                        ventas_por_ruta[nombre_linea_actual] = ventas_por_ruta.get(nombre_linea_actual, 0) + balance_float
                
                # Sumar a ventas de productos nuevos (IPN) - Lógica restaurada
                ciclo_vida = sale.get('product_life_cycle')
                if ciclo_vida and ciclo_vida == 'nuevo':
                    if nombre_linea_actual:
                        ventas_ipn_por_linea[nombre_linea_actual] = ventas_ipn_por_linea.get(nombre_linea_actual, 0) + balance_float
                
                # Agrupar por producto para Top 7
                producto_nombre = sale.get('name', '').strip()
                if producto_nombre:
                    ventas_por_producto[producto_nombre] = ventas_por_producto.get(producto_nombre, 0) + balance_float
                    if producto_nombre not in ciclo_vida_por_producto:
                        ciclo_vida_por_producto[producto_nombre] = ciclo_vida
                
                # Agrupar por ciclo de vida para el gráfico de dona
                ciclo_vida_grafico = ciclo_vida if ciclo_vida else 'No definido'
                ventas_por_ciclo_vida[ciclo_vida_grafico] = ventas_por_ciclo_vida.get(ciclo_vida_grafico, 0) + balance_float

                # Agrupar por forma farmacéutica para el gráfico de ECharts
                forma_farmaceutica = sale.get('pharmaceutical_forms_id')
                nombre_forma = forma_farmaceutica[1] if forma_farmaceutica and isinstance(forma_farmaceutica, list) and len(forma_farmaceutica) > 1 else 'Instrumental'
                ventas_por_forma[nombre_forma] = ventas_por_forma.get(nombre_forma, 0) + balance_float

        print(f"💰 Ventas por línea comercial: {ventas_por_linea}")
        print(f"📦 Ventas por Vencimiento (Ciclo de Vida): {ventas_por_ruta}")
        print(f"✨ Ventas IPN (Productos Nuevos): {ventas_ipn_por_linea}")

        # --- Procesamiento de datos para gráficos (después del bucle) ---

        # 1. Procesar datos para la tabla principal
        for linea in lineas_comerciales_estaticas:
            meta = metas_del_mes.get(linea['id'], 0)
            nombre_linea = linea['nombre'].upper()
            
            # Usar ventas reales de Odoo
            venta = ventas_por_linea.get(nombre_linea, 0)
            
            # Usar la meta IPN registrada por el usuario
            meta_pn = metas_ipn_del_mes.get(linea['id'], 0)
            venta_pn = ventas_ipn_por_linea.get(nombre_linea, 0) # Usar el cálculo real de ventas de productos nuevos
            vencimiento = ventas_por_ruta.get(nombre_linea, 0) # Usamos el nuevo cálculo
            
            porcentaje_total = (venta / meta * 100) if meta > 0 else 0
            porcentaje_pn = (venta_pn / meta_pn * 100) if meta_pn > 0 else 0

            datos_lineas.append({
                'nombre': linea['nombre'],
                'meta': meta,
                'venta': venta, # Ahora es positivo
                'porcentaje_total': (venta / meta * 100) if meta > 0 else 0,
                'meta_pn': meta_pn,
                'venta_pn': venta_pn,
                'porcentaje_pn': porcentaje_pn,
                'vencimiento_6_meses': vencimiento
            })
            
            total_meta += meta
            total_venta += venta
            total_meta_pn += meta_pn
            total_venta_pn += venta_pn
            total_vencimiento += vencimiento
        
        # 2. Calcular KPIs
        # Calcular KPIs
        kpis = {
            'meta_total': total_meta, # Meta siempre es positiva
            'venta_total': total_venta, # Ya es positivo
            'porcentaje_avance': (total_venta / total_meta * 100) if total_meta > 0 else 0,
            'meta_ipn': total_meta_pn, # Meta IPN siempre es positiva
            'venta_ipn': total_venta_pn, # Ya es positivo
            'porcentaje_avance_ipn': (total_venta_pn / total_meta_pn * 100) if total_meta_pn > 0 else 0,
            'vencimiento_6_meses': total_vencimiento,
            'avance_diario_total': ((total_venta / total_meta * 100) / dia_actual) if total_meta > 0 and dia_actual > 0 else 0,
            'avance_diario_ipn': ((total_venta_pn / total_meta_pn * 100) / dia_actual) if total_meta_pn > 0 and dia_actual > 0 else 0
        }
        
        # 3. Ordenar productos para el gráfico Top 7
        # Ordenar productos por ventas y tomar los top 7
        productos_ordenados = sorted(ventas_por_producto.items(), key=lambda x: x[1], reverse=True)[:7]
        
        datos_productos = []
        for nombre_producto, venta in productos_ordenados:
            datos_productos.append({
                'nombre': nombre_producto,
                'venta': venta,
                'ciclo_vida': ciclo_vida_por_producto.get(nombre_producto, 'No definido')
            })
        
        print(f"🏆 Top 7 productos por ventas: {[p['nombre'] for p in datos_productos]}")
        
        # 4. Ordenar datos para el gráfico de Ciclo de Vida
        # Convertir a lista ordenada por ventas
        datos_ciclo_vida = []
        for ciclo, venta in sorted(ventas_por_ciclo_vida.items(), key=lambda x: x[1], reverse=True):
            datos_ciclo_vida.append({
                'ciclo': ciclo,
                'venta': venta
            })
        
        print(f"📈 Ventas por Ciclo de Vida: {datos_ciclo_vida}")
        
        # 5. Ordenar datos para el gráfico de Forma Farmacéutica
        # Convertir a lista ordenada por ventas
        datos_forma_farmaceutica = []
        for forma, venta in sorted(ventas_por_forma.items(), key=lambda x: x[1], reverse=True):
            datos_forma_farmaceutica.append({
                'forma': forma,
                'venta': venta
            })
        
        return render_template('dashboard_clean.html',
                             meses_disponibles=meses_disponibles,
                             mes_seleccionado=mes_seleccionado,
                             mes_nombre=mes_nombre,
                             dia_actual=dia_actual,
                             kpis=kpis,
                             datos_lineas=datos_lineas, # Usar para gráficos
                             datos_lineas_tabla=datos_lineas, # Usar para la tabla
                             datos_productos=datos_productos,
                             datos_ciclo_vida=datos_ciclo_vida if 'datos_ciclo_vida' in locals() else [],
                             datos_forma_farmaceutica=datos_forma_farmaceutica)
    
    except Exception as e:
        flash(f'Error al obtener datos del dashboard: {str(e)}', 'danger')
        
        # Crear datos por defecto para evitar errores
        fecha_actual = datetime.now()
        kpis_default = {
            'meta_total': 0,
            'venta_total': 0,
            'porcentaje_avance': 0,
            'meta_ipn': 0,
            'venta_ipn': 0,
            'porcentaje_avance_ipn': 0,
            'vencimiento_6_meses': 0,
            'avance_diario_total': 0,
            'avance_diario_ipn': 0
        }
        
        return render_template('dashboard_clean.html',
                             meses_disponibles=[{
                                 'key': fecha_actual.strftime('%Y-%m'),
                                 'nombre': f"{fecha_actual.strftime('%B')} {fecha_actual.year}"
                             }],
                             mes_seleccionado=fecha_actual.strftime('%Y-%m'),
                             mes_nombre=f"{fecha_actual.strftime('%B').upper()} {fecha_actual.year}",
                             dia_actual=fecha_actual.day,
                             kpis=kpis_default,
                             datos_lineas=[], # Se mantiene vacío en caso de error
                             datos_lineas_tabla=[],
                             datos_productos=[],
                             datos_ciclo_vida=[],
                             datos_forma_farmaceutica=[])


@app.route('/dashboard_linea')
def dashboard_linea():
    if 'username' not in session:
        return redirect(url_for('login'))

    try:
        # --- 1. OBTENER FILTROS ---
        fecha_actual = datetime.now()
        mes_seleccionado = request.args.get('mes', fecha_actual.strftime('%Y-%m'))
        año_actual = fecha_actual.year
        meses_disponibles = get_meses_del_año(año_actual)

        linea_seleccionada_nombre = request.args.get('linea_nombre', 'PETMEDICA') # Default a PETMEDICA si no se especifica

        # Obtener día correcto según el mes seleccionado (lógica del dashboard principal)
        if mes_seleccionado == fecha_actual.strftime('%Y-%m'):
            # Mes actual: usar día actual
            dia_actual = fecha_actual.day
        else:
            # Mes pasado: usar último día del mes
            año_sel_dia, mes_sel_dia = mes_seleccionado.split('-')
            ultimo_dia_mes = calendar.monthrange(int(año_sel_dia), int(mes_sel_dia))[1]
            dia_actual = ultimo_dia_mes

        # Mapeo de nombre de línea a ID para cargar metas
        mapeo_nombre_a_id = {
            'PETMEDICA': 'petmedica', 'AGROVET': 'agrovet', 'PET NUTRISCIENCE': 'pet_nutriscience',
            'AVIVET': 'avivet', 'ECOMMERCE': 'ecommerce', 'OTROS': 'otros',
            'GENVET': 'genvet', 'LICITACIÓN': 'licitacion', 'Ninguno': 'ninguno'
        }
        linea_seleccionada_id = mapeo_nombre_a_id.get(linea_seleccionada_nombre.upper(), 'petmedica')

        # --- 2. OBTENER DATOS ---
        año_sel, mes_sel = mes_seleccionado.split('-')
        fecha_inicio = f"{año_sel}-{mes_sel}-01"
        ultimo_dia = calendar.monthrange(int(año_sel), int(mes_sel))[1]
        fecha_fin = f"{año_sel}-{mes_sel}-{ultimo_dia}"

        # Cargar metas de vendedores para el mes y línea seleccionados
        # La estructura es metas[equipo_id][vendedor_id][mes_key]
        metas_vendedores_historicas = LOCAL_STORAGE.get('metas_vendedores', {})
        # 1. Obtener todas las metas del equipo/línea
        metas_del_equipo = metas_vendedores_historicas.get(linea_seleccionada_id, {})

        # Obtener todos los vendedores de Odoo
        todos_los_vendedores = {str(v['id']): v['name'] for v in data_manager.get_all_sellers()}

        # Obtener ventas del mes
        sales_data = data_manager.get_sales_lines(
            date_from=fecha_inicio,
            date_to=fecha_fin,
            limit=10000
        )

        # --- 3. PROCESAR Y AGREGAR DATOS POR VENDEDOR ---
        ventas_por_vendedor = {}
        ventas_ipn_por_vendedor = {}
        ventas_vencimiento_por_vendedor = {}
        ventas_por_producto = {}
        ventas_por_ciclo_vida = {}
        ventas_por_forma = {}

        for sale in sales_data:
            linea_comercial = sale.get('commercial_line_national_id')
            if linea_comercial and isinstance(linea_comercial, list) and len(linea_comercial) > 1:
                nombre_linea_actual = linea_comercial[1].upper()

                # Filtrar por la línea comercial seleccionada
                if nombre_linea_actual == linea_seleccionada_nombre.upper():
                    user_info = sale.get('invoice_user_id')
                    if user_info and isinstance(user_info, list) and len(user_info) > 1:
                        vendedor_id = str(user_info[0])
                        balance = float(sale.get('balance', 0))

                        # Agrupar ventas totales
                        ventas_por_vendedor[vendedor_id] = ventas_por_vendedor.get(vendedor_id, 0) + balance

                        # Agrupar ventas IPN
                        if sale.get('product_life_cycle') == 'nuevo':
                            ventas_ipn_por_vendedor[vendedor_id] = ventas_ipn_por_vendedor.get(vendedor_id, 0) + balance

                        # Agrupar ventas por vencimiento < 6 meses
                        ruta = sale.get('route_id')
                        if isinstance(ruta, list) and len(ruta) > 0 and ruta[0] in [18, 19]:
                            ventas_vencimiento_por_vendedor[vendedor_id] = ventas_vencimiento_por_vendedor.get(vendedor_id, 0) + balance

                        # Agrupar para gráficos (Top Productos, Ciclo Vida, Forma Farmacéutica)
                        producto_nombre = sale.get('name', '').strip()
                        if producto_nombre:
                            ventas_por_producto[producto_nombre] = ventas_por_producto.get(producto_nombre, 0) + balance

                        ciclo_vida = sale.get('product_life_cycle', 'No definido')
                        ventas_por_ciclo_vida[ciclo_vida] = ventas_por_ciclo_vida.get(ciclo_vida, 0) + balance

                        forma_farma = sale.get('pharmaceutical_forms_id')
                        nombre_forma = forma_farma[1] if forma_farma and len(forma_farma) > 1 else 'Instrumental'
                        ventas_por_forma[nombre_forma] = ventas_por_forma.get(nombre_forma, 0) + balance

        # --- 4. CONSTRUIR ESTRUCTURA DE DATOS PARA LA PLANTILLA ---
        datos_vendedores = []
        total_meta = 0
        total_venta = 0
        total_meta_ipn = 0
        total_venta_ipn = 0
        total_vencimiento = 0

        # Iterar sobre todos los vendedores para incluirlos aunque no tengan ventas
        for vendedor_id, vendedor_nombre in todos_los_vendedores.items():
            # 2. Obtener la meta para este vendedor y este mes específico
            meta_guardada = metas_del_equipo.get(vendedor_id, {}).get(mes_seleccionado, {})
            
            meta = float(meta_guardada.get('meta', 0))
            meta_ipn = float(meta_guardada.get('meta_ipn', 0))
            venta = ventas_por_vendedor.get(vendedor_id, 0)
            venta_ipn = ventas_ipn_por_vendedor.get(vendedor_id, 0)
            vencimiento = ventas_vencimiento_por_vendedor.get(vendedor_id, 0)

            # Incluir solo si el vendedor tiene ventas en el mes seleccionado
            if venta > 0:
                datos_vendedores.append({
                    'id': vendedor_id,
                    'nombre': vendedor_nombre,
                    'meta': meta,
                    'venta': venta,
                    'porcentaje_avance': (venta / meta * 100) if meta > 0 else 0,
                    'meta_ipn': meta_ipn,
                    'venta_ipn': venta_ipn,
                    'porcentaje_avance_ipn': (venta_ipn / meta_ipn * 100) if meta_ipn > 0 else 0,
                    'vencimiento_6_meses': vencimiento
                })
                total_meta += meta
                total_venta += venta
                total_meta_ipn += meta_ipn
                total_venta_ipn += venta_ipn
                total_vencimiento += vencimiento

        # Ordenar por venta descendente
        datos_vendedores = sorted(datos_vendedores, key=lambda x: x['venta'], reverse=True)

        # KPIs generales para la línea
        kpis = {
            'meta_total': total_meta,
            'venta_total': total_venta,
            'porcentaje_avance': (total_venta / total_meta * 100) if total_meta > 0 else 0,
            'meta_ipn': total_meta_ipn,
            'venta_ipn': total_venta_ipn,
            'porcentaje_avance_ipn': (total_venta_ipn / total_meta_ipn * 100) if total_meta_ipn > 0 else 0,
            'vencimiento_6_meses': total_vencimiento,
            'avance_diario_total': ((total_venta / total_meta * 100) / dia_actual) if total_meta > 0 and dia_actual > 0 else 0,
            'avance_diario_ipn': ((total_venta_ipn / total_meta_ipn * 100) / dia_actual) if total_meta_ipn > 0 and dia_actual > 0 else 0
        }

        # Datos para gráficos
        productos_ordenados = sorted(ventas_por_producto.items(), key=lambda x: x[1], reverse=True)[:7]
        datos_productos = [{'nombre': n, 'venta': v} for n, v in productos_ordenados]

        datos_ciclo_vida = [{'ciclo': c, 'venta': v} for c, v in ventas_por_ciclo_vida.items()]
        datos_forma_farmaceutica = [{'forma': f, 'venta': v} for f, v in ventas_por_forma.items()]

        # Lista de todas las líneas para el selector
        lineas_comerciales_disponibles = [
            'PETMEDICA', 'AGROVET', 'PET NUTRISCIENCE', 'AVIVET', 'ECOMMERCE', 'OTROS', 'GENVET', 'LICITACIÓN'
        ]

        return render_template('dashboard_linea.html',
                               linea_nombre=linea_seleccionada_nombre,
                               mes_seleccionado=mes_seleccionado,
                               meses_disponibles=meses_disponibles,
                               kpis=kpis,
                               datos_vendedores=datos_vendedores,
                               datos_productos=datos_productos,
                               datos_ciclo_vida=datos_ciclo_vida,
                               datos_forma_farmaceutica=datos_forma_farmaceutica,
                               lineas_disponibles=lineas_comerciales_disponibles)

    except Exception as e:
        flash(f'Error al generar el dashboard para la línea: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))


@app.route('/meta', methods=['GET', 'POST'])
def meta():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    try:
        # Líneas comerciales estáticas de la empresa
        lineas_comerciales_estaticas = [
            {'nombre': 'PETMEDICA', 'id': 'petmedica'},
            {'nombre': 'AGROVET', 'id': 'agrovet'},
            {'nombre': 'PET NUTRISCIENCE', 'id': 'pet_nutriscience'},
            {'nombre': 'AVIVET', 'id': 'avivet'},
            {'nombre': 'ECOMMERCE', 'id': 'ecommerce'},
            {'nombre': 'OTROS', 'id': 'otros'},
            {'nombre': 'GENVET', 'id': 'genvet'},
            {'nombre': 'LICITACIÓN', 'id': 'licitacion'},
            {'nombre': 'Ninguno', 'id': 'ninguno'},
        ]
        
        # Obtener año actual y mes seleccionado
        fecha_actual = datetime.now()
        año_actual = fecha_actual.year
        mes_seleccionado = request.args.get('mes', fecha_actual.strftime('%Y-%m'))
        
        # Crear todos los meses del año actual
        meses_año = [{'es_actual': m['key'] == fecha_actual.strftime('%Y-%m'), **m} for m in get_meses_del_año(año_actual)]
        
        if request.method == 'POST':
            # Obtener el mes del formulario
            mes_formulario = request.form.get('mes_seleccionado', mes_seleccionado)
            
            # Procesar metas enviadas
            metas_data = {}
            metas_ipn_data = {}
            total_meta = 0
            total_meta_ipn = 0
            
            for linea in lineas_comerciales_estaticas:
                # Procesar Meta Total
                meta_value = request.form.get(f"meta_{linea['id']}", '0')
                try:
                    clean_value = str(meta_value).replace(',', '') if meta_value else '0'
                    valor = float(clean_value) if clean_value else 0.0
                    metas_data[linea['id']] = valor
                    total_meta += valor
                except (ValueError, TypeError):
                    metas_data[linea['id']] = 0.0
                
                # Procesar Meta IPN
                meta_ipn_value = request.form.get(f"meta_ipn_{linea['id']}", '0')
                try:
                    clean_value_ipn = str(meta_ipn_value).replace(',', '') if meta_ipn_value else '0'
                    valor_ipn = float(clean_value_ipn) if clean_value_ipn else 0.0
                    metas_ipn_data[linea['id']] = valor_ipn
                    total_meta_ipn += valor_ipn
                except (ValueError, TypeError):
                    metas_ipn_data[linea['id']] = 0.0
            
            # Encontrar el nombre del mes
            mes_obj = next((m for m in meses_año if m['key'] == mes_formulario), None)
            mes_nombre_formulario = mes_obj['nombre'] if mes_obj else ""
            
            metas_historicas = LOCAL_STORAGE.get('metas_por_linea', {})
            metas_historicas[mes_formulario] = {
                'metas': metas_data,
                'metas_ipn': metas_ipn_data,
                'total': total_meta,
                'total_ipn': total_meta_ipn,
                'mes_nombre': mes_nombre_formulario
            }
            LOCAL_STORAGE['metas_por_linea'] = metas_historicas
            
            flash(f'Metas guardadas exitosamente para {mes_nombre_formulario}. Total: S/ {total_meta:,.0f}', 'success')
            
            # Actualizar mes seleccionado después de guardar
            mes_seleccionado = mes_formulario
        
        # Obtener todas las metas históricas
        metas_historicas = LOCAL_STORAGE.get('metas_por_linea', {})
        
        # Obtener metas y total del mes seleccionado
        metas_actuales = metas_historicas.get(mes_seleccionado, {}).get('metas', {})
        metas_ipn_actuales = metas_historicas.get(mes_seleccionado, {}).get('metas_ipn', {})
        total_actual = sum(metas_actuales.values()) if metas_actuales else 0
        total_ipn_actual = sum(metas_ipn_actuales.values()) if metas_ipn_actuales else 0
        
        # Encontrar el nombre del mes seleccionado
        mes_obj_seleccionado = next((m for m in meses_año if m['key'] == mes_seleccionado), meses_año[fecha_actual.month - 1])
        
        return render_template('meta.html',
                             lineas_comerciales=lineas_comerciales_estaticas,
                             metas_actuales=metas_actuales,
                             metas_ipn_actuales=metas_ipn_actuales,
                             metas_historicas=metas_historicas,
                             meses_año=meses_año,
                             mes_seleccionado=mes_seleccionado,
                             mes_nombre=mes_obj_seleccionado['nombre'],
                             total_actual=total_actual,
                             total_ipn_actual=total_ipn_actual,
                             fecha_actual=fecha_actual)
    
    except Exception as e:
        flash(f'Error al procesar metas: {str(e)}', 'danger')
        return render_template('meta.html',
                             lineas_comerciales=[],
                             metas_actuales={},
                             metas_ipn_actuales={},
                             metas_historicas={},
                             meses_año=[],
                             mes_seleccionado="",
                             mes_nombre="",
                             total_actual=0,
                             total_ipn_actual=0,
                             fecha_actual=datetime.now())

# --- API RUTAS PARA COBRANZA INTERNACIONAL ---
@app.route('/api/cobranza_internacional/kpis')
def api_cobranza_internacional_kpis():
    if 'username' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        date_from = request.args.get('start')
        date_to = request.args.get('end')
        payment_state = request.args.get('payment_state')
        linea_id = request.args.get('linea_id')
        
        kpis = data_manager.cobranza.get_cobranza_kpis_internacional(
            date_from, date_to, payment_state, linea_id
        )
        
        return jsonify(kpis)
    
    except Exception as e:
        print(f"[ERROR] api_cobranza_internacional_kpis: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/cobranza_internacional/top15')
def api_cobranza_internacional_top15():
    if 'username' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        date_from = request.args.get('start')
        date_to = request.args.get('end')
        
        top15 = data_manager.cobranza.get_top15_deudores_internacional(date_from, date_to)
        
        return jsonify(top15)
    
    except Exception as e:
        print(f"[ERROR] api_cobranza_internacional_top15: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/cobranza_internacional/aging')
def api_cobranza_internacional_aging():
    if 'username' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        date_from = request.args.get('start')
        date_to = request.args.get('end')
        
        kpis = data_manager.cobranza.get_cobranza_kpis_internacional(date_from, date_to)
        
        # Formatear para gráfico
        aging_buckets = kpis.get('aging_buckets', {})
        return jsonify({
            'labels': ['Vigente', '1-30 días', '31-60 días', '61-90 días', '+90 días'],
            'values': [
                aging_buckets.get('vigente', 0),
                aging_buckets.get('1-30', 0),
                aging_buckets.get('31-60', 0),
                aging_buckets.get('61-90', 0),
                aging_buckets.get('+90', 0),
            ]
        })
    
    except Exception as e:
        print(f"[ERROR] api_cobranza_internacional_aging: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/cobranza_internacional/dso_by_country')
def api_cobranza_internacional_dso_by_country():
    if 'username' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        date_from = request.args.get('start')
        date_to = request.args.get('end')
        
        kpis = data_manager.cobranza.get_cobranza_kpis_internacional(date_from, date_to)
        dso_by_country = kpis.get('dso_by_country', {})
        
        return jsonify({
            'countries': list(dso_by_country.keys()),
            'dso_values': list(dso_by_country.values())
        })
    
    except Exception as e:
        print(f"[ERROR] api_cobranza_internacional_dso_by_country: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/cobranza_internacional/dso_trend')
def api_cobranza_internacional_dso_trend():
    if 'username' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Placeholder - se necesitaría implementar cálculo por mes
        return jsonify({
            'labels': ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun'],
            'dso_values': [45, 48, 52, 49, 53, 51],
            'objetivo': [45, 45, 45, 45, 45, 45]
        })
    
    except Exception as e:
        print(f"[ERROR] api_cobranza_internacional_dso_trend: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/export/excel/sales')
def export_excel_sales():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    try:
        # Obtener filtros de la URL
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        linea_id = request.args.get('linea_id')
        partner_id = request.args.get('partner_id')
        
        # Convertir a tipos apropiados
        if linea_id:
            try:
                linea_id = int(linea_id)
            except (ValueError, TypeError):
                linea_id = None
        
        if partner_id:
            try:
                partner_id = int(partner_id)
            except (ValueError, TypeError):
                partner_id = None
        
        # Obtener datos
        sales_data = data_manager.get_sales_lines(
            date_from=date_from,
            date_to=date_to,
            partner_id=partner_id,
            linea_id=linea_id,
            limit=10000  # Más datos para export
        )
        
        # Filtrar VENTA INTERNACIONAL (exportaciones)
        sales_data_filtered = []
        for sale in sales_data:
            linea_comercial = sale.get('commercial_line_national_id')
            if linea_comercial and isinstance(linea_comercial, list) and len(linea_comercial) > 1:
                nombre_linea = linea_comercial[1].upper()
                if 'VENTA INTERNACIONAL' in nombre_linea:
                    continue
            
            # También filtrar por canal de ventas
            canal_ventas = sale.get('sales_channel_id')
            if canal_ventas and isinstance(canal_ventas, list) and len(canal_ventas) > 1:
                nombre_canal = canal_ventas[1].upper()
                if 'VENTA INTERNACIONAL' in nombre_canal or 'INTERNACIONAL' in nombre_canal:
                    continue
            
            sales_data_filtered.append(sale)
        
        # Crear DataFrame
        df = pd.DataFrame(sales_data_filtered)
        
        # Crear archivo Excel en memoria
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Ventas', index=False)
        
        output.seek(0)
        
        # Generar nombre de archivo con timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'ventas_farmaceuticas_{timestamp}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Error al exportar datos: {str(e)}', 'danger')
        return redirect(url_for('sales'))

@app.route('/metas_vendedor', methods=['GET', 'POST'])
def metas_vendedor():
    if 'username' not in session:
        return redirect(url_for('login'))

    # Obtener meses y líneas comerciales para los filtros
    fecha_actual = datetime.now()
    año_actual = fecha_actual.year
    meses_disponibles = get_meses_del_año(año_actual)
    lineas_comerciales_estaticas = [
        {'nombre': 'PETMEDICA', 'id': 'petmedica'},
        {'nombre': 'AGROVET', 'id': 'agrovet'},
        {'nombre': 'PET NUTRISCIENCE', 'id': 'pet_nutriscience'},
        {'nombre': 'AVIVET', 'id': 'avivet'},
        {'nombre': 'ECOMMERCE', 'id': 'ecommerce'},
        {'nombre': 'OTROS', 'id': 'otros'},
        {'nombre': 'GENVET', 'id': 'genvet'},
        {'nombre': 'LICITACIÓN', 'id': 'licitacion'},
        {'nombre': 'Ninguno', 'id': 'ninguno'},
    ]
    equipos_definidos = [
        {'id': 'petmedica', 'nombre': 'PETMEDICA'},
        {'id': 'agrovet', 'nombre': 'AGROVET'},
        {'id': 'pet_nutriscience', 'nombre': 'PET NUTRISCIENCE'},
        {'id': 'avivet', 'nombre': 'AVIVET'},
        {'id': 'ecommerce', 'nombre': 'ECOMMERCE'},
        {'id': 'otros', 'nombre': 'OTROS'},
    ]

    # Determinar mes y línea seleccionados (desde form o por defecto)
    mes_seleccionado = request.form.get('mes_seleccionado', fecha_actual.strftime('%Y-%m'))
    linea_seleccionada = request.form.get('linea_seleccionada', lineas_comerciales_estaticas[0]['id'])

    if request.method == 'POST':
        # --- 1. GUARDAR ASIGNACIONES DE EQUIPOS ---
        equipo_actualizado_id = request.form.get('guardar_equipo') # Para el mensaje flash
        todos_los_vendedores_para_guardar = data_manager.get_all_sellers()
        equipos_guardados = LOCAL_STORAGE.get('equipos', {})

        for equipo in equipos_definidos:
            campo_vendedores = f'vendedores_{equipo["id"]}'
            if campo_vendedores in request.form:
                vendedores_str = request.form.get(campo_vendedores, '')
                if vendedores_str:
                    vendedores_ids = [int(vid) for vid in vendedores_str.split(',') if vid.isdigit()]
                    equipos_guardados[equipo['id']] = vendedores_ids
                else:
                    equipos_guardados[equipo['id']] = []
        LOCAL_STORAGE['equipos'] = equipos_guardados

        # --- 2. GUARDAR TODAS LAS METAS (ESTRUCTURA PIVOT) ---
        metas_vendedores_historicas = LOCAL_STORAGE.get('metas_vendedores', {})
        
        for equipo in equipos_definidos:
            equipo_id = equipo['id']
            if equipo_id not in metas_vendedores_historicas:
                metas_vendedores_historicas[equipo_id] = {}

            vendedores_ids_en_equipo = equipos_guardados.get(equipo_id, [])
            for vendedor_id in vendedores_ids_en_equipo:
                vendedor_id_str = str(vendedor_id)
                if vendedor_id_str not in metas_vendedores_historicas[equipo_id]:
                    metas_vendedores_historicas[equipo_id][vendedor_id_str] = {}

                for mes in meses_disponibles:
                    mes_key = mes['key']
                    # No es necesario crear la clave del mes aquí, se crea si hay datos

                    meta_valor_str = request.form.get(f'meta_{equipo_id}_{vendedor_id_str}_{mes_key}')
                    meta_ipn_valor_str = request.form.get(f'meta_ipn_{equipo_id}_{vendedor_id_str}_{mes_key}')

                    # Convertir a float, manejar valores vacíos como None para no guardar ceros innecesarios
                    meta = float(meta_valor_str) if meta_valor_str else None
                    meta_ipn = float(meta_ipn_valor_str) if meta_ipn_valor_str else None

                    if meta is not None or meta_ipn is not None:
                        # Si la clave del mes no existe, créala
                        if mes_key not in metas_vendedores_historicas[equipo_id][vendedor_id_str]:
                             metas_vendedores_historicas[equipo_id][vendedor_id_str][mes_key] = {}
                        metas_vendedores_historicas[equipo_id][vendedor_id_str][mes_key] = {
                            'meta': meta or 0.0,
                            'meta_ipn': meta_ipn or 0.0
                        }
                    # Si ambos son None y la clave existe, se elimina para limpiar el JSON
                    elif mes_key in metas_vendedores_historicas[equipo_id][vendedor_id_str]:
                        del metas_vendedores_historicas[equipo_id][vendedor_id_str][mes_key]

        LOCAL_STORAGE['metas_vendedores'] = metas_vendedores_historicas
        
        if equipo_actualizado_id:
            flash(f'Miembros del equipo actualizados. Ahora puedes asignar sus metas.', 'info')
        else:
            flash('Equipos y metas guardados correctamente.', 'success')

        # Redirigir con los parámetros para recargar la página con los filtros correctos
        return redirect(url_for('metas_vendedor'))

    # GET o después de POST
    todos_los_vendedores = data_manager.get_all_sellers()
    vendedores_por_id = {v['id']: v for v in todos_los_vendedores}
    equipos_guardados = LOCAL_STORAGE.get('equipos', {})

    # Construir la estructura de datos para la plantilla
    equipos_con_vendedores = []
    for equipo_def in equipos_definidos:
        equipo_id = equipo_def['id']
        vendedores_ids = equipos_guardados.get(equipo_id, [])
        vendedores_de_equipo = [vendedores_por_id[vid] for vid in vendedores_ids if vid in vendedores_por_id]
        
        equipos_con_vendedores.append({
            'id': equipo_id,
            'nombre': equipo_def['nombre'],
            'vendedores_ids': [str(vid) for vid in vendedores_ids], # Para Tom-Select
            'vendedores': sorted(vendedores_de_equipo, key=lambda v: v['name']) # Para la tabla
        })

    # Para la vista, pasamos todas las metas cargadas
    metas_guardadas = LOCAL_STORAGE.get('metas_vendedores', {})

    return render_template('metas_vendedor.html',
                           meses_disponibles=meses_disponibles,
                           lineas_comerciales=lineas_comerciales_estaticas,
                           equipos_con_vendedores=equipos_con_vendedores,
                           todos_los_vendedores=todos_los_vendedores,
                           metas_guardadas=metas_guardadas)

@app.route('/export/dashboard/details')
def export_dashboard_details():
    if 'username' not in session:
        return redirect(url_for('login'))

    try:
        # Obtener el mes seleccionado de los parámetros de la URL
        mes_seleccionado = request.args.get('mes')
        if not mes_seleccionado:
            flash('No se especificó un mes para la exportación.', 'danger')
            return redirect(url_for('dashboard'))

        # Calcular fechas para el mes seleccionado
        año_sel, mes_sel = mes_seleccionado.split('-')
        fecha_inicio = f"{año_sel}-{mes_sel}-01"
        ultimo_dia = calendar.monthrange(int(año_sel), int(mes_sel))[1]
        fecha_fin = f"{año_sel}-{mes_sel}-{ultimo_dia}"

        # Obtener datos de ventas reales desde Odoo para ese mes
        sales_data = data_manager.get_sales_lines(
            date_from=fecha_inicio,
            date_to=fecha_fin,
            limit=10000  # Límite alto para exportación
        )

        # Filtrar VENTA INTERNACIONAL (exportaciones), igual que en el dashboard
        sales_data_filtered = []
        for sale in sales_data:
            linea_comercial = sale.get('commercial_line_national_id')
            if linea_comercial and isinstance(linea_comercial, list) and len(linea_comercial) > 1:
                if 'VENTA INTERNACIONAL' in linea_comercial[1].upper():
                    continue
            
            canal_ventas = sale.get('sales_channel_id')
            if canal_ventas and isinstance(canal_ventas, list) and len(canal_ventas) > 1:
                nombre_canal = canal_ventas[1].upper()
                if 'VENTA INTERNACIONAL' in nombre_canal or 'INTERNACIONAL' in nombre_canal:
                    continue
            
            sales_data_filtered.append(sale)

        # Convertir el balance a positivo para que coincida con el dashboard
        for sale in sales_data_filtered:
            if 'balance' in sale and sale['balance'] is not None:
                sale['balance'] = float(sale['balance']) # Ya viene con el signo correcto desde OdooManager

        # Crear DataFrame de Pandas con los datos filtrados
        df = pd.DataFrame(sales_data_filtered)

        # Crear archivo Excel en memoria
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=f'Detalle Ventas {mes_seleccionado}', index=False)
        output.seek(0)

        # Generar nombre de archivo
        filename = f'detalle_ventas_{mes_seleccionado}.xlsx'

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        flash(f'Error al exportar los detalles del dashboard: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))


# --- NUEVA RUTA PARA REPORTE CxC GENERAL ---
@app.route('/reporte_cxc_general', methods=['GET', 'POST'])
def reporte_cxc_general():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    try:
        # Obtener filtros de la request
        selected_filters = {
            'date_from': request.form.get('date_from') or request.args.get('date_from'),
            'date_to': request.form.get('date_to') or request.args.get('date_to'),
            'customer': request.form.get('customer') or request.args.get('customer'),
            'account_codes': request.form.get('account_codes') or request.args.get('account_codes'),
            'search_term': request.form.get('search_term') or request.args.get('search_term')
        }
        
        # Convertir strings vacíos a None
        for key, value in selected_filters.items():
            if value == '':
                selected_filters[key] = None
        
        # Obtener datos de CxC usando el método del bi_creditos_cobranzas
        cxc_data = data_manager.get_report_lines(
            start_date=selected_filters['date_from'],
            end_date=selected_filters['date_to'],
            customer=selected_filters['customer'],
            account_codes=selected_filters['account_codes'],
            search_term=selected_filters['search_term'],
            limit=1000
        )
        
        return render_template('reporte_cxc_general.html', 
                             cxc_data=cxc_data,
                             selected_filters=selected_filters,
                             fecha_actual=datetime.now())
    
    except Exception as e:
        flash(f'Error al obtener datos de CxC: {str(e)}', 'danger')
        return render_template('reporte_cxc_general.html', 
                             cxc_data=[],
                             selected_filters={},
                             fecha_actual=datetime.now())

# --- NUEVA RUTA PARA REPORTE INTERNACIONAL ---
@app.route('/reporte_internacional', methods=['GET', 'POST'])
def reporte_internacional():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    try:
        # Obtener filtros
        selected_filters = {
            'date_from': request.form.get('date_from') or request.args.get('date_from'),
            'date_to': request.form.get('date_to') or request.args.get('date_to'),
            'customer': request.form.get('customer') or request.args.get('customer'),
            'payment_state': request.form.get('payment_state') or request.args.get('payment_state')
        }
        
        # Convertir strings vacíos a None
        for key, value in selected_filters.items():
            if value == '':
                selected_filters[key] = None
        
        # Obtener datos internacional con campos calculados
        internacional_data = data_manager.get_report_internacional(
            start_date=selected_filters['date_from'],
            end_date=selected_filters['date_to'],
            customer=selected_filters['customer'],
            payment_state=selected_filters['payment_state'],
            limit=2000
        )
        
        # Ordenar por días vencidos (descendente - mayor a menor)
        if internacional_data:
            internacional_data = sorted(
                internacional_data, 
                key=lambda x: x.get('dias_vencido', 0), 
                reverse=True
            )
        
        return render_template('reporte_internacional.html',
                             internacional_data=internacional_data,
                             selected_filters=selected_filters,
                             fecha_actual=datetime.now())
    
    except Exception as e:
        flash(f'Error al obtener datos internacionales: {str(e)}', 'danger')
        return render_template('reporte_internacional.html',
                             internacional_data=[],
                             selected_filters={},
                             fecha_actual=datetime.now())

# --- NUEVA RUTA PARA DASHBOARD COBRANZA INTERNACIONAL ---
@app.route('/dashboard_cobranza_internacional')
def dashboard_cobranza_internacional():
    if 'username' not in session:
        return redirect(url_for('login'))
    return render_template('dashboard_cobranza_internacional.html')

# --- API RUTAS PARA COBRANZA ---
@app.route('/api/cobranza/kpis')
def api_cobranza_kpis():
    if 'username' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Obtener filtros
        date_from = request.args.get('start')
        date_to = request.args.get('end')
        payment_state = request.args.get('payment_state')
        
        # Obtener datos de cobranza
        kpis_data = data_manager.get_cobranza_kpis(date_from, date_to, payment_state)
        
        return jsonify(kpis_data)
    
    except Exception as e:
        print(f"Error en api_cobranza_kpis: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/cobranza/top15')
def api_cobranza_top15():
    if 'username' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Obtener filtros
        date_from = request.args.get('start')
        date_to = request.args.get('end')
        payment_state = request.args.get('payment_state')
        
        # Obtener top 15 clientes
        top15_data = data_manager.get_top15_cobranza(date_from, date_to, payment_state)
        
        return jsonify(top15_data)
    
    except Exception as e:
        print(f"Error en api_cobranza_top15: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/cobranza/top15/details')
def api_cobranza_top15_details():
    if 'username' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Obtener filtros
        date_from = request.args.get('start')
        date_to = request.args.get('end')
        payment_state = request.args.get('payment_state')
        
        # Obtener detalles del top 15
        details_data = data_manager.get_top15_cobranza_details(date_from, date_to, payment_state)
        
        return jsonify(details_data)
    
    except Exception as e:
        print(f"Error en api_cobranza_top15_details: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/cobranza/lineas')
def api_cobranza_lineas():
    if 'username' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Obtener líneas comerciales usando el método de filtros
        filter_options = data_manager.get_filter_options()
        lineas = filter_options.get('lineas', [])
        
        # Formatear para la respuesta JSON
        lineas_data = [{'id': l['id'], 'name': l['display_name']} for l in lineas]
        return jsonify(lineas_data)
    
    except Exception as e:
        print(f"Error en api_cobranza_lineas: {e}")
        return jsonify([]), 200  # Retornar lista vacía en lugar de error 500

@app.route('/api/cobranza/linea')
def api_cobranza_linea():
    if 'username' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    
    try:
        # Obtener filtros
        date_from = request.args.get('start')
        date_to = request.args.get('end')
        payment_state = request.args.get('payment_state')
        linea_id = request.args.get('linea_id')
        
        # Verificar si el método existe
        if hasattr(data_manager, 'get_cobranza_por_linea'):
            linea_data = data_manager.get_cobranza_por_linea(date_from, date_to, payment_state, linea_id)
        else:
            # Retornar estructura vacía si el método no existe
            linea_data = {'rows': []}
        
        return jsonify(linea_data)
    
    except Exception as e:
        print(f"Error obteniendo cobranza por línea: {e}")
        return jsonify({'rows': []}), 200  # Retornar estructura vacía en lugar de error

# --- RUTA DE EXPORTACIÓN PARA CxC ---
@app.route('/export/excel/cxc')
def export_excel_cxc():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    try:
        # Obtener filtros de la URL
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        customer = request.args.get('customer')
        account_codes = request.args.get('account_codes')
        
        # Obtener datos de CxC
        cxc_data = data_manager.get_report_lines(
            start_date=date_from,
            end_date=date_to,
            customer=customer,
            account_codes=account_codes,
            limit=10000  # Más datos para export
        )
        
        if not cxc_data:
            flash('No hay datos para exportar con los filtros seleccionados.', 'warning')
            return redirect(url_for('reporte_cxc_general'))
        
        # Mapeo de nombres técnicos a nombres en español (como en el frontend)
        column_mapping = {
            'payment_state': 'Estado de Pago',
            'invoice_date': 'Fecha de Factura',
            'I10nn_latam_document_type_id': 'Tipo de Documento',
            'move_name': 'Número de Factura',
            'invoice_origin': 'Origen',
            'account_id/code': 'Código de Cuenta',
            'account_id/name': 'Nombre de Cuenta',
            'patner_id/vat': 'RUC/DNI',
            'patner_id': 'Cliente',
            'currency_id': 'Moneda',
            'amount_total': 'Monto Total',
            'amount_residual': 'Importe Adeudado',
            'invoice_date_due': 'Fecha de Vencimiento',
            'ref': 'Referencia',
            'invoice_payment_term_id': 'Condición de Pago',
            'name': 'Descripción',
            'move_id/invoice_user_id': 'Vendedor',
            'patner_id/state_id': 'Provincia',
            'patner_id/l10n_pe_district': 'Distrito',
            'patner_id/country_code': 'Código de País',
            'patner_id/country_id': 'País',
            'sub_channel_id': 'Sub Canal',
            'move_id/sales_channel_id': 'Canal de Venta',
            'move_id/sales_type_id': 'Tipo de Venta'
        }
        
        # Crear DataFrame con solo las columnas que queremos
        df = pd.DataFrame(cxc_data)
        
        # Reordenar y renombrar columnas según el frontend
        ordered_columns = [
            'invoice_date', 'I10nn_latam_document_type_id', 'move_name', 'invoice_origin',
            'account_id/code', 'account_id/name', 'patner_id/vat', 'patner_id', 'currency_id',
            'amount_total', 'amount_residual', 'invoice_date_due', 'ref', 
            'invoice_payment_term_id', 'name', 'move_id/invoice_user_id',
            'patner_id/state_id', 'patner_id/l10n_pe_district', 'patner_id/country_code',
            'patner_id/country_id', 'sub_channel_id', 'move_id/sales_channel_id', 'move_id/sales_type_id'
        ]
        
        # Filtrar solo columnas que existen
        existing_columns = [col for col in ordered_columns if col in df.columns]
        df = df[existing_columns]
        
        # Renombrar columnas al español
        df.rename(columns=column_mapping, inplace=True)
        
        # Crear archivo Excel con formato profesional
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Cuentas por Cobrar', index=False, startrow=1)
            
            # Obtener el workbook y worksheet
            workbook = writer.book
            worksheet = writer.sheets['Cuentas por Cobrar']
            
            # Importar módulos de estilo de openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Estilos
            header_fill = PatternFill(start_color="875A7B", end_color="875A7B", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            border_thin = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # Título del reporte
            worksheet.merge_cells('A1:W1')
            title_cell = worksheet['A1']
            title_cell.value = 'REPORTE DE CUENTAS POR COBRAR - CUENTA 12'
            title_cell.font = Font(size=14, bold=True, color="875A7B")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Aplicar estilos a los encabezados
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=2, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border_thin
            
            # Ajustar ancho de columnas y formato de celdas
            column_widths = {
                'Fecha de Factura': 15,
                'Tipo de Documento': 18,
                'Número de Factura': 18,
                'Origen': 20,
                'Código de Cuenta': 15,
                'Nombre de Cuenta': 25,
                'RUC/DNI': 15,
                'Cliente': 30,
                'Moneda': 12,
                'Monto Total': 15,
                'Importe Adeudado': 18,
                'Fecha de Vencimiento': 18,
                'Referencia': 20,
                'Condición de Pago': 20,
                'Descripción': 35,
                'Vendedor': 25,
                'Provincia': 20,
                'Distrito': 20,
                'Código de País': 15,
                'País': 15,
                'Sub Canal': 18,
                'Canal de Venta': 20,
                'Tipo de Venta': 18
            }
            
            # Aplicar anchos y formatos
            for col_num, column_title in enumerate(df.columns, 1):
                col_letter = get_column_letter(col_num)
                
                # Ajustar ancho
                width = column_widths.get(column_title, 15)
                worksheet.column_dimensions[col_letter].width = width
                
                # Aplicar formato a las celdas de datos
                for row_num in range(3, len(df) + 3):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.border = border_thin
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    
                    # Formato específico por tipo de columna
                    if column_title in ['Monto Total', 'Importe Adeudado']:
                        # Formato de moneda
                        cell.number_format = 'S/ #,##0.00'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    elif column_title in ['Fecha de Factura', 'Fecha de Vencimiento']:
                        # Formato de fecha
                        cell.number_format = 'DD/MM/YYYY'
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        # Texto general
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Agregar filtros automáticos
            worksheet.auto_filter.ref = worksheet.dimensions
            
            # Congelar la fila de encabezados
            worksheet.freeze_panes = 'A3'
        
        output.seek(0)
        
        # Generar nombre de archivo con timestamp y filtros
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filters_suffix = ""
        if date_from or date_to:
            filters_suffix = f"_{date_from or 'inicio'}_{date_to or 'hoy'}"
        
        filename = f'reporte_cxc_general{filters_suffix}_{timestamp}.xlsx'
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        flash(f'Error al exportar datos de CxC: {str(e)}', 'danger')
        return redirect(url_for('reporte_cxc_general'))

@app.route('/export/excel/internacional')
def export_excel_internacional():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    try:
        # Obtener filtros
        date_from = request.args.get('date_from')
        date_to = request.args.get('date_to')
        customer = request.args.get('customer')
        payment_state = request.args.get('payment_state')
        
        # Obtener datos
        internacional_data = data_manager.get_report_internacional(
            start_date=date_from,
            end_date=date_to,
            customer=customer,
            payment_state=payment_state,
            limit=10000
        )
        
        if not internacional_data:
            flash('No hay datos para exportar con los filtros seleccionados.', 'warning')
            return redirect(url_for('reporte_internacional'))
        
        # Ordenar por días vencidos (descendente)
        internacional_data = sorted(internacional_data, key=lambda x: x.get('dias_vencido', 0), reverse=True)
        
        # Mapeo de columnas
        column_mapping = {
            'payment_state': 'Estado de Pago',
            'vat': 'Cod. Extranjero',
            'patner_id': 'Cliente',
            'I10nn_latam_document_type_id': 'Tipo de Documento',
            'name': 'Factura',
            'invoice_origin': 'Origen',
            'invoice_payment_term_id': 'Condicion de Pago',
            'invoice_date': 'Fecha de Factura',
            'invoice_date_due': 'Fecha de Vencimiento',
            'currency_id': 'Moneda',
            'amount_total_currency_signed': 'Total USD',
            'amount_residual_with_retention': 'Adeudado USD',
            'monto_interes': 'Monto de Interes',
            'dias_vencido': 'Dias de Vencido',
            'estado_deuda': 'Estado de Deuda',
            'antiguedad': 'Antiguedad',
            'invoice_user_id': 'Vendedor',
            'country_code': 'Codigo de Pais',
            'country_id': 'Pais'
        }
        
        df = pd.DataFrame(internacional_data)
        
        # Reordenar y renombrar
        ordered_columns = [
            'payment_state', 'vat', 'patner_id', 'I10nn_latam_document_type_id',
            'name', 'invoice_origin', 'invoice_payment_term_id', 'invoice_date',
            'invoice_date_due', 'currency_id', 'amount_total_currency_signed',
            'amount_residual_with_retention', 'monto_interes', 'dias_vencido',
            'estado_deuda', 'antiguedad', 'invoice_user_id', 'country_code', 'country_id'
        ]
        
        df = df[[col for col in ordered_columns if col in df.columns]]
        df = df.rename(columns=column_mapping)
        
        # Crear Excel con formato profesional usando openpyxl
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Reporte Internacional', index=False, startrow=1)
            
            workbook = writer.book
            worksheet = writer.sheets['Reporte Internacional']
            
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Estilos (igual que reporte_cxc_general)
            header_fill = PatternFill(start_color="875A7B", end_color="875A7B", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            border_thin = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            
            # Título
            worksheet.merge_cells('A1:S1')
            title_cell = worksheet['A1']
            title_cell.value = 'REPORTE INTERNACIONAL - FACTURAS NO PAGADAS'
            title_cell.font = Font(size=14, bold=True, color="875A7B")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Aplicar estilos a encabezados
            for col_num, column_title in enumerate(df.columns, 1):
                cell = worksheet.cell(row=2, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border_thin
            
            # Ajustar anchos y formatos de celdas
            for col_num, column_title in enumerate(df.columns, 1):
                col_letter = get_column_letter(col_num)
                worksheet.column_dimensions[col_letter].width = 18
                
                for row_num in range(3, len(df) + 3):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.border = border_thin
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    
                    # Formato específico por tipo de columna
                    if 'USD' in column_title or 'Interes' in column_title:
                        cell.number_format = '$ #,##0.00'
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    elif 'Fecha' in column_title:
                        cell.number_format = 'DD/MM/YYYY'
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Agregar filtros automáticos
            worksheet.auto_filter.ref = worksheet.dimensions
            
            # Congelar la fila de encabezados
            worksheet.freeze_panes = 'A3'
        
        output.seek(0)
        
        # Nombre del archivo con fecha
        filename = f"reporte_internacional_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error exportando internacional: {e}")
        flash(f'Error al exportar datos: {str(e)}', 'danger')
        return redirect(url_for('reporte_internacional'))

if __name__ == '__main__':
    print("[INFO] Iniciando Dashboard de Cobranzas...")
    print("[INFO] Disponible en: http://127.0.0.1:5002")
    print("[INFO] Usuario: configurado en .env")
    print("[INFO] Debug: ON")
    app.run(debug=True, port=5002)
